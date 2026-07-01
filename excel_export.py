from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Iterable, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from quote_core import AnalysisResult, BatchAnalysisResult, QuoteRow


QuoteSource = AnalysisResult | BatchAnalysisResult


def _iter_rows(source: QuoteSource) -> Iterable[Tuple[str, str, QuoteRow]]:
    if isinstance(source, AnalysisResult):
        for row in source.quote_rows:
            yield source.source_file, "；".join(source.warnings), row
        return
    for item in source.items:
        if not item.result:
            continue
        for row in item.result.quote_rows:
            yield item.source_file, "；".join(item.result.warnings), row


def _style_header(ws, row: int, fill: str = "D9EAF7") -> None:
    thin = Side(style="thin", color="B7C3D0")
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)


def _autosize(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col].width = min(max(max_len + 2, 10), 28)


def write_quote_xlsx(source: QuoteSource, out_path: str | Path) -> None:
    rows = list(_iter_rows(source))
    wb = Workbook()
    quote_ws = wb.active
    quote_ws.title = "正式报价单"
    detail_ws = wb.create_sheet("内部核算明细")

    quote_ws.merge_cells("A1:J1")
    quote_ws["A1"] = "激光切割报价单"
    quote_ws["A1"].font = Font(size=16, bold=True)
    quote_ws["A1"].alignment = Alignment(horizontal="center")
    quote_ws["A2"] = "报价日期"
    quote_ws["B2"] = date.today().isoformat()

    quote_headers = ["序号", "图号", "名称", "材质", "厚度(mm)", "规格(mm)", "数量", "单价", "金额", "备注"]
    quote_ws.append([])
    quote_ws.append(quote_headers)
    _style_header(quote_ws, 4)

    total = 0.0
    for idx, (_, warning, row) in enumerate(rows, start=1):
        total += row.amount
        note_parts = [row.note] if row.note else []
        if warning:
            note_parts.append(warning)
        quote_ws.append([
            idx,
            row.drawing_no,
            row.name,
            row.material,
            row.thickness_mm,
            row.size_mm,
            row.quantity,
            round(row.unit_price, 4),
            round(row.amount, 4),
            "；".join(note_parts),
        ])
    total_row = quote_ws.max_row + 1
    quote_ws.cell(total_row, 8, "合计")
    quote_ws.cell(total_row, 9, round(total, 4))
    quote_ws.cell(total_row, 8).font = Font(bold=True)
    quote_ws.cell(total_row, 9).font = Font(bold=True)

    detail_headers = [
        "来源文件", "图号", "名称", "零件序号", "重复视图数", "材质", "厚度(mm)", "数量",
        "尺寸(mm)", "孔数", "穿孔数", "切割米数", "毛面积(mm2)", "净面积(mm2)",
        "毛重(kg)", "净重(kg)", "材料费", "切割费", "穿孔费", "废料抵扣",
        "其他工序费", "基础单价", "单价", "金额", "备注",
    ]
    detail_ws.append(detail_headers)
    _style_header(detail_ws, 1, fill="E8F3E8")
    for source_file, warning, row in rows:
        note_parts = [row.note] if row.note else []
        if warning:
            note_parts.append(warning)
        detail_ws.append([
            source_file,
            row.drawing_no,
            row.name,
            row.part_index,
            row.duplicate_count,
            row.material,
            row.thickness_mm,
            row.quantity,
            row.size_mm,
            row.hole_count,
            row.pierce_count,
            round(row.cut_length_m, 4),
            round(row.gross_area_mm2, 4),
            round(row.net_area_mm2, 4),
            round(row.gross_weight_kg, 4),
            round(row.net_weight_kg, 4),
            round(row.material_fee_each, 2),
            round(row.cut_fee_each, 2),
            round(row.pierce_fee_each, 2),
            round(row.scrap_credit_each, 2),
            round(row.other_process_fee_each, 2),
            round(row.base_unit_price, 4),
            round(row.unit_price, 4),
            round(row.amount, 4),
            "；".join(note_parts),
        ])

    for ws in (quote_ws, detail_ws):
        ws.freeze_panes = "A2"
        _autosize(ws)
    quote_ws.freeze_panes = "A5"
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
